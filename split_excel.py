import openpyxl
from openpyxl.styles import PatternFill, Font
import os

source_file = os.path.join("Test Results", "Agriculture", "Farmer_Buyer_Connect_Test_Cases.xlsx")
farmer_file = os.path.join("Test Results", "Agriculture", "Farmer_Test_Cases.xlsx")
buyer_file = os.path.join("Test Results", "Agriculture", "Buyer_Test_Cases.xlsx")

wb_source = openpyxl.load_workbook(source_file)
ws_source = wb_source["All Test Cases"]

# Create new workbooks
wb_farmer = openpyxl.Workbook()
ws_farmer = wb_farmer.active
ws_farmer.title = "Farmer Test Cases"

wb_buyer = openpyxl.Workbook()
ws_buyer = wb_buyer.active
ws_buyer.title = "Buyer Test Cases"

# Copy headers
headers = []
role_col_idx = None
status_col_idx = None

for col in range(1, ws_source.max_column + 1):
    cell_val = ws_source.cell(row=1, column=col).value
    headers.append(cell_val)
    ws_farmer.cell(row=1, column=col, value=cell_val)
    ws_buyer.cell(row=1, column=col, value=cell_val)
    if cell_val == "User Role":
        role_col_idx = col
    if cell_val == "Status":
        status_col_idx = col

# Freeze panes
ws_farmer.freeze_panes = "A2"
ws_buyer.freeze_panes = "A2"

# Filter and copy rows
farmer_row_idx = 2
buyer_row_idx = 2

pass_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
pass_font = Font(color="006100", bold=True)

for row in range(2, ws_source.max_row + 1):
    role = ws_source.cell(row=row, column=role_col_idx).value
    if role == "Farmer":
        target_ws = ws_farmer
        target_row = farmer_row_idx
        farmer_row_idx += 1
    elif role == "Buyer":
        target_ws = ws_buyer
        target_row = buyer_row_idx
        buyer_row_idx += 1
    else:
        continue # Skip any other roles just in case

    for col in range(1, ws_source.max_column + 1):
        cell_val = ws_source.cell(row=row, column=col).value
        target_cell = target_ws.cell(row=target_row, column=col, value=cell_val)
        
        # Apply green coloring if it's the status column
        if col == status_col_idx and cell_val == "PASS":
            target_cell.fill = pass_fill
            target_cell.font = pass_font

wb_farmer.save(farmer_file)
wb_buyer.save(buyer_file)
print("Successfully split the test cases into separate files!")
