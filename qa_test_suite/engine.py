"""Engine for creating the QA Test Suite Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
GRAY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def style_data(ws, rows, cols):
    for r in range(2, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            cell.font = Font(name='Calibri', size=10)

def add_conditional_formatting(ws, col_letter, max_row):
    rng = f'{col_letter}2:{col_letter}{max_row}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Pass"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Fail"'], fill=RED_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Blocked"'], fill=YELLOW_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Not Executed"'], fill=GRAY_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"N/A"'], fill=BLUE_FILL))

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_test_cases_sheet(wb, test_cases):
    ws = wb.active
    ws.title = 'Test_Cases'
    headers = ['Test Case ID','Module','Sub Module','Test Type','Test Scenario',
        'Test Case Description','Preconditions','Test Data','Test Steps',
        'Expected Result','Actual Result','Priority','Severity','Environment',
        'Browser/Device','Automation Tool','Automation Status','API/UI/DB',
        'Execution Status','Pass/Fail','Defect ID','Defect Description',
        'Evidence','Tester','Execution Date','Remarks']
    ws.append(headers)
    for tc in test_cases:
        ws.append(tc)
    style_header(ws, len(headers))
    style_data(ws, len(test_cases)+1, len(headers))
    add_conditional_formatting(ws, 'T', len(test_cases)+1)
    widths = [12,18,18,14,30,35,25,20,35,30,28,10,10,12,14,14,14,10,14,12,10,25,14,12,14,20]
    set_col_widths(ws, widths)
    return ws

def create_summary_sheet(wb, test_cases, modules):
    ws = wb.create_sheet('Test_Summary')
    ws.append(['Metric','Value'])
    tc_count = len(test_cases)
    r = tc_count + 1
    ref = f"Test_Cases!T2:T{r}"
    ws.append(['Total Test Cases', tc_count])
    ws.append(['Executed', f'=COUNTIF({ref},"Pass")+COUNTIF({ref},"Fail")'])
    ws.append(['Passed', f'=COUNTIF({ref},"Pass")'])
    ws.append(['Failed', f'=COUNTIF({ref},"Fail")'])
    ws.append(['Blocked', f'=COUNTIF({ref},"Blocked")'])
    ws.append(['Not Executed', f'=COUNTIF({ref},"Not Executed")'])
    ws.append(['Not Applicable', f'=COUNTIF({ref},"N/A")'])
    ws.append(['Pass %', f'=IF(B3=0,0,ROUND(B4/B3*100,2))'])
    ws.append(['Fail %', f'=IF(B3=0,0,ROUND(B5/B3*100,2))'])
    ws.append(['Blocked %', f'=IF(B2=0,0,ROUND(B6/B2*100,2))'])
    ws.append([])
    ws.append(['--- Module-wise Breakdown ---',''])
    sev_ref = f"Test_Cases!M2:M{r}"
    mod_ref = f"Test_Cases!B2:B{r}"
    ws.append(['Module','Total','Pass','Fail'])
    for m in modules:
        row_n = ws.max_row + 1
        ws.append([m,
            f'=COUNTIF({mod_ref},A{row_n})',
            f'=COUNTIFS({mod_ref},A{row_n},{ref},"Pass")',
            f'=COUNTIFS({mod_ref},A{row_n},{ref},"Fail")'])
    ws.append([])
    ws.append(['--- Defect Summary ---',''])
    ws.append(['Critical Defects', f'=COUNTIF({sev_ref},"Critical")'])
    ws.append(['High Defects', f'=COUNTIF({sev_ref},"High")'])
    ws.append(['Medium Defects', f'=COUNTIF({sev_ref},"Medium")'])
    ws.append(['Low Defects', f'=COUNTIF({sev_ref},"Low")'])
    style_header(ws, 4)
    style_data(ws, ws.max_row, 4)
    set_col_widths(ws, [30, 15, 15, 15])

def create_defect_log(wb):
    ws = wb.create_sheet('Defect_Log')
    headers = ['Defect ID','Test Case ID','Module','Defect Title','Description',
        'Steps to Reproduce','Expected Result','Actual Result','Severity',
        'Priority','Environment','Browser/Device','Status','Assigned To',
        'Created Date','Fixed Date','Retest Result','Remarks']
    ws.append(headers)
    ws.append(['--','--','--','No defects logged - execution pending','Actual application testing required',
        'N/A','N/A','N/A','--','--','--','--','Open','Unassigned',
        datetime.now().strftime('%Y-%m-%d'),'--','Pending','Awaiting test execution'])
    style_header(ws, len(headers))
    style_data(ws, 2, len(headers))
    set_col_widths(ws, [12,12,16,25,30,30,25,25,10,10,12,14,10,14,14,14,12,20])

def create_automation_matrix(wb, test_cases):
    ws = wb.create_sheet('Automation_Matrix')
    headers = ['Test Case ID','Module','Automation Candidate','Selenium/Appium/API',
        'Automation Priority','Locator/Endpoint','Framework','Script Status',
        'Execution Status','Remarks']
    ws.append(headers)
    for tc in test_cases:
        tid, mod = tc[0], tc[1]
        layer = tc[17]
        if layer == 'UI':
            tool = 'Selenium'
            loc = 'CSS/ID Locators'
            fw = 'Selenium+Java+TestNG+POM'
        elif layer == 'API':
            tool = 'REST Assured'
            loc = 'API Endpoint'
            fw = 'REST Assured+Java+TestNG'
        elif layer == 'DB':
            tool = 'JDBC'
            loc = 'SQL Query'
            fw = 'JDBC+Java+TestNG'
        else:
            tool = 'Selenium'
            loc = 'CSS/ID Locators'
            fw = 'Selenium+Java+TestNG+POM'
        cand = 'Yes' if tc[16] in ['Automatable','Automated'] else 'No'
        prio = tc[11]
        ws.append([tid, mod, cand, tool, prio, loc, fw, 'Not Started', 'Not Executed', ''])
    style_header(ws, len(headers))
    style_data(ws, len(test_cases)+1, len(headers))
    set_col_widths(ws, [12,18,18,16,16,20,28,14,14,20])

def create_performance_sheet(wb):
    ws = wb.create_sheet('Performance_Results')
    headers = ['Test ID','Scenario','Concurrent Users','Requests',
        'Avg Response Time(ms)','Min Response Time(ms)','Max Response Time(ms)',
        'Throughput(req/s)','Error Rate(%)','CPU Usage(%)','Memory Usage(%)',
        'Network Usage(MB)','Result','Remarks']
    ws.append(headers)
    scenarios = [
        ('PT-001','Homepage Load',10,100),('PT-002','Homepage Load',50,500),
        ('PT-003','Homepage Load',100,1000),('PT-004','Login API',10,100),
        ('PT-005','Login API',50,500),('PT-006','Login API',100,1000),
        ('PT-007','Crop Search',10,100),('PT-008','Crop Search',50,500),
        ('PT-009','Crop Search',100,1000),('PT-010','Crop Listing',50,500),
        ('PT-011','Order Creation',10,100),('PT-012','Order Creation',50,500),
        ('PT-013','Requirement Post',50,500),('PT-014','Dashboard Load',100,1000),
        ('PT-015','Concurrent Login',250,2500),('PT-016','Stress Test - Search',500,5000),
        ('PT-017','Spike Test - Orders',1000,5000),('PT-018','Endurance - Browse',100,10000),
        ('PT-019','Volume - Crop DB',50,5000),('PT-020','Scalability - Full',5000,25000),
    ]
    for s in scenarios:
        ws.append([s[0],s[1],s[2],s[3],'Not Executed','Not Executed','Not Executed',
            'Not Executed','Not Executed','Not Executed','Not Executed','Not Executed',
            'Not Executed','Requires JMeter execution'])
    style_header(ws, len(headers))
    style_data(ws, len(scenarios)+1, len(headers))
    set_col_widths(ws, [10,22,16,12,18,18,18,16,12,12,14,14,12,24])

def create_security_sheet(wb):
    ws = wb.create_sheet('Security_Testing')
    headers = ['Security Test ID','Vulnerability Category','Test Scenario',
        'Test Steps','Expected Result','Actual Result','Severity','Status',
        'Evidence','Remediation','Retest Status']
    ws.append(headers)
    sec_tests = [
        ('SEC-001','SQL Injection','SQL injection on login email field','Enter " OR 1=1-- in email field and submit','Input rejected/sanitized'),
        ('SEC-002','SQL Injection','SQL injection on search field','Enter UNION SELECT in crop search','Query parameterized, no data leak'),
        ('SEC-003','XSS','Stored XSS on crop description','Enter <script>alert(1)</script> in description','Script tags sanitized/escaped'),
        ('SEC-004','XSS','Reflected XSS via URL params','Add <img onerror=alert(1)> in URL param','Input encoded in response'),
        ('SEC-005','CSRF','CSRF on order creation','Submit order form without CSRF token','Request rejected with 403'),
        ('SEC-006','CSRF','CSRF on profile update','Craft cross-site request to update profile','Anti-CSRF token validated'),
        ('SEC-007','Broken Auth','Access dashboard without login','Navigate directly to /dashboard','Redirected to login page'),
        ('SEC-008','Broken Auth','Use expired JWT token','Send API request with expired token','401 Unauthorized returned'),
        ('SEC-009','IDOR','Access other farmer profile by ID','Change farmer_id in URL/API','403 Forbidden returned'),
        ('SEC-010','IDOR','View other buyer orders','Modify order_id in request','Access denied'),
        ('SEC-011','Privilege Escalation','Buyer accesses farmer endpoints','Call farmer API with buyer token','403 Forbidden'),
        ('SEC-012','Privilege Escalation','User accesses admin panel','Navigate to /admin as regular user','Access denied'),
        ('SEC-013','Session Hijacking','Reuse stolen session token','Copy session cookie to new browser','Session invalidated/IP check'),
        ('SEC-014','Session Fixation','Set session before login','Set JSESSIONID before authentication','New session created on login'),
        ('SEC-015','JWT Manipulation','Modify JWT payload','Change role claim in JWT','Signature validation fails'),
        ('SEC-016','JWT Manipulation','Use none algorithm','Set JWT alg to none','Token rejected'),
        ('SEC-017','Password Security','Brute force login','Attempt 20 wrong passwords','Account locked after 5 attempts'),
        ('SEC-018','Rate Limiting','API flood test','Send 100 requests/second to login','Rate limit 429 returned'),
        ('SEC-019','File Upload','Upload malicious PHP file','Upload .php file as crop image','File type rejected'),
        ('SEC-020','File Upload','Upload oversized file','Upload 50MB image','Size limit enforced'),
        ('SEC-021','Path Traversal','Access system files','Request ../../etc/passwd in URL','Path traversal blocked'),
        ('SEC-022','Info Exposure','Check error messages','Trigger server error','Generic error, no stack trace'),
        ('SEC-023','API Auth','Access API without token','Call protected endpoint without auth header','401 returned'),
        ('SEC-024','Security Headers','Verify response headers','Check X-Frame-Options, CSP, HSTS','All security headers present'),
        ('SEC-025','HTTPS','Test HTTP to HTTPS redirect','Access site via HTTP','Redirected to HTTPS'),
        ('SEC-026','CORS','Test cross-origin request','Send request from unauthorized origin','CORS policy blocks request'),
        ('SEC-027','Cookie Security','Check cookie flags','Inspect session cookies','HttpOnly, Secure, SameSite set'),
        ('SEC-028','Input Validation','Test all forms for script injection','Enter script tags in all text inputs','All inputs sanitized'),
    ]
    for s in sec_tests:
        ws.append([s[0],s[1],s[2],s[3],s[4],'Not Executed - Requires security tool execution',
            'High' if 'Injection' in s[1] or 'XSS' in s[1] else 'Medium',
            'Not Executed','Not Available','Pending assessment','Not Tested'])
    style_header(ws, len(headers))
    style_data(ws, len(sec_tests)+1, len(headers))
    set_col_widths(ws, [14,20,28,35,28,32,10,14,14,22,14])

def create_dashboard(wb, test_cases, modules):
    ws = wb.create_sheet('Dashboard')
    tc_count = len(test_cases)
    r = tc_count + 1
    ref = f"Test_Cases!T2:T{r}"
    ws.append(['SMART AGRICULTURE QA DASHBOARD'])
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(['Metric','Count'])
    ws.append(['Total Test Cases', tc_count])
    ws.append(['Passed', f'=COUNTIF({ref},"Pass")'])
    ws.append(['Failed', f'=COUNTIF({ref},"Fail")'])
    ws.append(['Blocked', f'=COUNTIF({ref},"Blocked")'])
    ws.append(['Not Executed', f'=COUNTIF({ref},"Not Executed")'])
    ws.append(['Not Applicable', f'=COUNTIF({ref},"N/A")'])
    ws.append(['Pass %', f'=IF(B4+B5=0,0,ROUND(B5/(B5+B6)*100,2))'])
    ws.append(['Fail %', f'=IF(B4+B5=0,0,ROUND(B6/(B5+B6)*100,2))'])
    sev = f"Test_Cases!M2:M{r}"
    ws.append([])
    ws.append(['Critical Defects', f'=COUNTIFS({sev},"Critical",{ref},"Fail")'])
    ws.append(['High Defects', f'=COUNTIFS({sev},"High",{ref},"Fail")'])
    ws.append(['Medium Defects', f'=COUNTIFS({sev},"Medium",{ref},"Fail")'])
    ws.append(['Low Defects', f'=COUNTIFS({sev},"Low",{ref},"Fail")'])
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = Font(name='Calibri', size=11)
    for row in [3]:
        for c in range(1,3):
            ws.cell(row=row,column=c).font = HEADER_FONT
            ws.cell(row=row,column=c).fill = HEADER_FILL
    set_col_widths(ws, [22, 16, 4, 22, 16, 16])
    # Chart 1: Execution Status Pie
    pie = PieChart()
    pie.title = "Execution Status"
    pie.style = 10
    cats = Reference(ws, min_col=1, min_row=5, max_row=9)
    vals = Reference(ws, min_col=2, min_row=5, max_row=9)
    pie.add_data(vals, titles_from_data=False)
    pie.set_categories(cats)
    pie.width = 16
    pie.height = 12
    ws.add_chart(pie, "D3")
    # Chart 2: Module breakdown
    mod_ref = f"Test_Cases!B2:B{r}"
    ws_row = ws.max_row + 2
    ws.cell(row=ws_row, column=1, value='Module').font = HEADER_FONT
    ws.cell(row=ws_row, column=1).fill = HEADER_FILL
    ws.cell(row=ws_row, column=2, value='Count').font = HEADER_FONT
    ws.cell(row=ws_row, column=2).fill = HEADER_FILL
    start_r = ws_row + 1
    for m in modules:
        ws.append([m, f'=COUNTIF({mod_ref},A{ws.max_row+1})'])
    end_r = ws.max_row
    bar = BarChart()
    bar.title = "Test Cases by Module"
    bar.style = 10
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Module"
    cats2 = Reference(ws, min_col=1, min_row=start_r, max_row=end_r)
    vals2 = Reference(ws, min_col=2, min_row=start_r, max_row=end_r)
    bar.add_data(vals2, titles_from_data=False)
    bar.set_categories(cats2)
    bar.width = 28
    bar.height = 14
    bar.shape = 4
    ws.add_chart(bar, "D20")
    # Chart 3: Defects by severity
    pie2 = PieChart()
    pie2.title = "Defects by Severity"
    pie2.style = 10
    cats3 = Reference(ws, min_col=1, min_row=13, max_row=16)
    vals3 = Reference(ws, min_col=2, min_row=13, max_row=16)
    pie2.add_data(vals3, titles_from_data=False)
    pie2.set_categories(cats3)
    pie2.width = 16
    pie2.height = 12
    ws.add_chart(pie2, "D40")
