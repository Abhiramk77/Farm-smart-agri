import openpyxl
from openpyxl.styles import PatternFill, Font
import sys
import os

excel_path = os.path.join("Test Results", "Agriculture", "Farmer_Buyer_Connect_Test_Cases.xlsx")

if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path)
pass_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
pass_font = Font(color="006100", bold=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Find the Status column index
    status_col_idx = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value == "Status":
            status_col_idx = col
            break
            
    if status_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col_idx)
            if cell.value == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font

wb.save(excel_path)
print("Successfully colored PASS cells green!")
